"""
The field should look like this:

       col0   col1   col2    col3   col4   col5    col6   col7   col8
    ||======|======|======||======|======|======||======|======|======||
    ||  A   |  A   |  A   ||  B   |  B   |  B   ||  C   |  C   |  C   ||
row0||cell0 |cell1 |cell2 ||cell3 |cell4 |cell5 ||cell6 |cell7 |cell8 ||
    ||______|______|______||______|______|______||______|______|______||
    ||  A   |  A   |  A   ||  B   |  B   |  B   ||  C   |  C   |  C   ||
row1||cell9 |cell10|cell11||cell12|cell13|cell14||cell15|cell16|cell17||
    ||______|______|______||______|______|______||______|______|______||
    ||  A   |  A   |  A   ||  B   |  B   |  B   ||  C   |  C   |  C   ||
row2||cell18|cell19|cell20||cell21|cell22|cell23||cell24|cell25|cell26||
    ||======|======|======||======|======|======||======|======|======||
    ||  D   |  D   |  D   ||  E   |  E   |  E   ||  F   |  F   |  F   ||
row3||cell27|cell28|cell29||cell30|cell31|cell32||cell33|cell34|cell35||
    ||______|______|______||______|______|______||______|______|______||
    ||  D   |  D   |  D   ||  E   |  E   |  E   ||  F   |  F   |  F   ||
row4||cell36|cell37|cell38||cell39|cell40|cell41||cell42|cell43|cell44||
    ||______|______|______||______|______|______||______|______|______||
    ||  D   |  D   |  D   ||  E   |  E   |  E   ||  F   |  F   |  F   ||
row5||cell45|cell46|cell47||cell48|cell49|cell50||cell51|cell52|cell53||
    ||======|======|======||======|======|======||======|======|======||
    ||  G   |  G   |  G   ||  H   |  H   |  H   ||  I   |  I   |  I   ||
row6||cell54|cell55|cell56||cell57|cell58|cell59||cell60|cell61|cell62||
    ||______|______|______||______|______|______||______|______|______||
    ||  G   |  G   |  G   ||  H   |  H   |  H   ||  I   |  I   |  I   ||
row7||cell63|cell64|cell65||cell66|cell67|cell68||cell69|cell70|cell71||
    ||______|______|______||______|______|______||______|______|______||
    ||  G   |  G   |  G   ||  H   |  H   |  H   ||  I   |  I   |  I   ||
row8||cell72|cell73|cell74||cell75|cell76|cell77||cell78|cell79|cell80||
    ||======|======|======||======|======|======||======|======|======||
"""
import openpyxl
DIGITS = (1, 2, 3, 4, 5, 6, 7, 8, 9)
rows = []
cols = []
squares = []
cells = {}


class Cell:
    def __init__(self, row, col, value=''):
        self.possible_values = list(DIGITS)
        self.value = value
        self.isSolved = False
        self.row = row
        self.col = col


class CellGroup:
    def __init__(self):
        self.cells = []
        self.possible_values = list(DIGITS)


def init_structure():
    global rows
    global cols
    global squares
    global cells
    # Initialize empty rows, cols and squares
    for index in range(0, 9):
        rows.append(CellGroup())
        cols.append(CellGroup())
        squares.append(CellGroup())
    # Initialize empty cells
    for cell_index in range(0, 81):
        cell_name = f'cell{cell_index}'
        row_index = cell_index // 9
        col_index = cell_index % 9
        # Create cell from class
        cells[cell_name] = Cell(row_index, col_index)
        # Adding it to a row and cols list
        rows[row_index].cells.append(cells[cell_name])
        cols[col_index].cells.append(cells[cell_name])
        # Adding squares
        # Maybe someday something shorter and not that straightforward?
        if row_index < 3:
            if col_index < 3:
                squares[0].cells.append(cells[cell_name])
            elif 3 <= col_index < 6:
                squares[1].cells.append(cells[cell_name])
            elif 6 <= col_index < 9:
                squares[2].cells.append(cells[cell_name])
        elif 3 <= row_index < 6:
            if col_index < 3:
                squares[3].cells.append(cells[cell_name])
            elif 3 <= col_index < 6:
                squares[4].cells.append(cells[cell_name])
            elif 6 <= col_index < 9:
                squares[5].cells.append(cells[cell_name])
        elif 6 <= row_index < 9:
            if col_index < 3:
                squares[6].cells.append(cells[cell_name])
            elif 3 <= col_index < 6:
                squares[7].cells.append(cells[cell_name])
            elif 6 <= col_index < 9:
                squares[8].cells.append(cells[cell_name])


def read_puzzle_xls():
    """Read initial know values from Excel"""
    global cells
    exlw = openpyxl.load_workbook('sudoku.xlsx',
                                  read_only=True,
                                  data_only=True)
    ws = exlw.active
    for cell_name, cell in cells.items():
        excel_cell_row = cell.row+1
        excel_cell_column = cell.col+1
        if ws.cell(excel_cell_row, excel_cell_column).value is not None:
            cell.value = ws.cell(
                excel_cell_row, excel_cell_column
            ).value
            cell.possible_values.clear()


def sanity_check():
    """If cell has value - clear possible_values"""
    global cells
    for cell_name, cell in cells.items():
        if cell.value is not None:
            cell.possible_values.clear()


def solve_group(group):
    for cell in group.cells:
        # Removing know values from line possible values
        if (cell.value != '') and (cell.value in group.possible_values):
            group.possible_values.remove(cell.value)
    # Removing line impossible values from cell possible values
    for cell in group.cells:
        for cell_pv in cell.possible_values:
            if cell_pv not in group.possible_values:
                cell.possible_values.remove(cell_pv)
        # Set value if only 1 possible value available
        if len(cell.possible_values) == 1:
            cell.value = cell.possible_values.pop()


def print_puzzle_debug():
    """Prints puzzle results and debug"""
    global cells
    global rows
    for row in rows:
        OutputLine = ''
        for cell in row.cells:
            OutputLine += str(cell.value) + ' '
        print(OutputLine)


def solve_puzzle():
    """Main program to solve the puzzle"""
    # Geeting all the cells without value to list
    unresolved_cells = []
    for cell_name, cell in cells.items():
        if cell.value == '':
            unresolved_cells.append(cell_name)
    # Solving only unknown cells
    for cell_name in unresolved_cells:
        # Solving groups
        for group in rows + cols + squares:
            solve_group(group)


init_structure()
read_puzzle_xls()
solve_puzzle()
print_puzzle_debug()
